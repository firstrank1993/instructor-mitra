import sys
import json
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def thin_border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

def medium_border():
    m = Side(style='medium')
    return Border(left=m, right=m, top=m, bottom=m)

def rand_int(a, b):
    return random.randint(a, b)

def distribute_marks(target_mark, is_out_of_30):
    target_60 = round(target_mark * 2) if is_out_of_30 else round(target_mark * 6)
    clamped = max(20, min(60, target_60))

    b = c = d = e = f = 0
    attempts = 0
    while attempts < 500:
        b = rand_int(2, 5)
        c = rand_int(2, 5)
        d = rand_int(3, 9)
        remaining = clamped - b - c - d
        if remaining < 16 or remaining > 40:
            attempts += 1
            continue
        e_min = max(8, remaining - 20)
        e_max = min(20, remaining - 8)
        if e_min > e_max:
            attempts += 1
            continue
        e = rand_int(e_min, e_max)
        f = remaining - e
        if f < 8 or f > 20 or e == f:
            attempts += 1
            continue
        break

    if attempts >= 500:
        b, c, d = 3, 3, 6
        rem = clamped - b - c - d
        e = min(20, max(8, rem // 2))
        f = rem - e
        if f < 8:
            e = rem - 8
            f = 8
        if f > 20:
            e = rem - 20
            f = 20

    total_60 = b + c + d + e + f
    return b, c, d, e, f, total_60


def style(cell, value='', bold=False, size=11, h='left', v='center',
          wrap=False, border=False):
    cell.value = value
    cell.font = Font(name='Calibri', size=size, bold=bold)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    if border:
        cell.border = thin_border()


def generate_subject_report(data, output_path):
    subject_type = data.get('subjectType', 'ES')
    trainees = data['trainees']
    subject_marks = data.get('subjectMarks', [])
    instructor = data['instructorData']
    batch = data['batchData']
    half = data['half']
    assessment_date = data.get('assessmentDate', '')
    trade = data.get('tradeData', {}) or {}
    has_5_subjects = data.get('has5Subjects', False)

    is_out_of_30 = (not has_5_subjects) and (subject_type == 'ES')

    title_map = {
        'ES': 'ANNEXURE-III (FAR-2 )',
        'WCS': '(FAR-2 )',
        'ED': '(FAR-2 )',
    }
    subject_title_map = {
        'ES': 'FORMAT FOR INTERNAL ASSESSMENT FOR EMPLOYABILITY SKILLS',
        'WCS': 'FORMAT FOR INTERNAL ASSESSMENT FOR WORKSHOP CALCULATION & SCIENCE',
        'ED': 'FORMAT FOR INTERNAL ASSESSMENT FOR ENGINEERING DRAWING',
    }
    convert_label = (
        f'Convert Total Marks in  to 30 Markes =\n{{(Col.G)/2}}'
        if is_out_of_30
        else f'Convert Total Marks in  to 10 Markes =\n{{(Col.G)/6}}'
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{subject_type}_Report'

    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    # Exact column widths from actual report
    widths = {1: 4.83, 2: 20.0, 3: 23.66, 4: 7.83, 5: 15.16,
              6: 7.5, 7: 8.0, 8: 7.33, 9: 8.0, 10: 8.43, 11: 15.16, 12: 8.16}
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Outer border: every used cell gets a border drawn manually at the end
    # ROW 1: report title (A1:L1)
    ws.row_dimensions[1].height = 17.85
    ws.merge_cells('A1:L1')
    style(ws.cell(row=1, column=1), title_map[subject_type], bold=True, size=14, h='center')

    # ROW 2: Internal Assessment (A2:L2)
    ws.row_dimensions[2].height = 18
    ws.merge_cells('A2:L2')
    style(ws.cell(row=2, column=1), 'Internal Assessment', bold=True, size=12, h='center')

    # ROW 3: subject title (A3:L3)
    ws.row_dimensions[3].height = 18
    ws.merge_cells('A3:L3')
    style(ws.cell(row=3, column=1), subject_title_map[subject_type], bold=True, size=10, h='center')

    # ROW 4: Assessor / Year of Enrolment
    ws.row_dimensions[4].height = 18
    ws.merge_cells('A4:C4')
    style(ws.cell(row=4, column=1), 'Name & Adddress of the Assessor')
    ws.merge_cells('D4:F4')
    style(ws.cell(row=4, column=4), instructor.get('displayName', ''), bold=True)
    ws.merge_cells('G4:J4')
    style(ws.cell(row=4, column=7), 'Year of Enrolment')
    ws.merge_cells('K4:L4')
    style(ws.cell(row=4, column=11), batch.get('yearOfAssessment', ''), bold=True)

    # ROW 5: ITI / Date of Assessment
    ws.row_dimensions[5].height = 18
    ws.merge_cells('A5:C5')
    style(ws.cell(row=5, column=1), 'Name & Address of ITI (Govt/Pvt)')
    ws.merge_cells('D5:F5')
    style(ws.cell(row=5, column=4), instructor.get('itiName', ''), bold=True)
    ws.merge_cells('G5:J5')
    style(ws.cell(row=5, column=7), 'Date of Assessment')
    ws.merge_cells('K5:L5')
    style(ws.cell(row=5, column=11), assessment_date, bold=True)

    # ROW 6: Industry / Assessment Location
    ws.row_dimensions[6].height = 18
    ws.merge_cells('A6:C6')
    style(ws.cell(row=6, column=1), 'Name & Address of the Industry')
    ws.merge_cells('D6:F6')
    style(ws.cell(row=6, column=4), instructor.get('address', ''), bold=True)
    ws.merge_cells('G6:J6')
    style(ws.cell(row=6, column=7), 'Assessment Location')
    ws.merge_cells('K6:L6')
    style(ws.cell(row=6, column=11), instructor.get('itiName', ''), bold=True)

    # ROW 7: Trade Name / Duration / SEM
    ws.row_dimensions[7].height = 24
    ws.merge_cells('A7:B7')
    style(ws.cell(row=7, column=1), 'Trade Name')
    ws.merge_cells('C7:F7')
    style(ws.cell(row=7, column=3), trade.get('name', ''), bold=True, h='center')
    ws.merge_cells('G7:I7')
    style(ws.cell(row=7, column=7), 'Duration Of  Trade')
    duration = trade.get('duration', 1)
    style(ws.cell(row=7, column=10), f"{duration} Year", bold=True)
    style(ws.cell(row=7, column=11), 'SEM')
    style(ws.cell(row=7, column=12), half, bold=True)

    # ROW 8: Learning Outcome / Batch NO
    ws.row_dimensions[8].height = 18
    ws.merge_cells('A8:F8')
    style(ws.cell(row=8, column=1), 'Learning Outcome :')
    ws.merge_cells('G8:J8')
    style(ws.cell(row=8, column=7), 'Batch NO')
    ws.merge_cells('K8:L8')
    style(ws.cell(row=8, column=11), batch.get('batchNumber', ''), bold=True)

    # ROW 9: Column headers
    ws.row_dimensions[9].height = 63.6
    style(ws.cell(row=9, column=1), 'Roll\nNo', bold=True, h='left', wrap=True, border=True)
    ws.merge_cells('B9:C9')
    style(ws.cell(row=9, column=2), 'Name', bold=True, h='center', border=True)
    style(ws.cell(row=9, column=3), '', border=True)
    style(ws.cell(row=9, column=4), 'Attendance', bold=True, h='center', v='center', wrap=True, border=True)
    style(ws.cell(row=9, column=5),
          'Speed for WC & Sc\n/ Accuracy of ED /\nComminacation\nskill fro ES',
          bold=True, h='center', wrap=True, border=True)
    ws.merge_cells('F9:G9')
    style(ws.cell(row=9, column=6),
          'Creative Work\n(Chart , Model\n,Poster , Project\nwork etc..)',
          bold=True, h='center', wrap=True, border=True)
    style(ws.cell(row=9, column=7), '', border=True)
    style(ws.cell(row=9, column=8), 'Quarterly -1', bold=True, h='center', wrap=True, border=True)
    style(ws.cell(row=9, column=9), 'Quarterly -2', bold=True, h='center', wrap=True, border=True)
    style(ws.cell(row=9, column=10), 'Total', bold=True, h='center', wrap=True, border=True)
    style(ws.cell(row=9, column=11), convert_label, bold=True, h='center', wrap=True, border=True)
    style(ws.cell(row=9, column=12), 'Sign of\nTrainee', bold=True, h='center', wrap=True, border=True)

    # ROW 10: Maximum Marks
    ws.row_dimensions[10].height = 18
    style(ws.cell(row=10, column=1), 'Maximum Marks =>', bold=True, border=True)
    ws.merge_cells('B10:C10')
    style(ws.cell(row=10, column=2), '', border=True)
    style(ws.cell(row=10, column=3), '', border=True)
    style(ws.cell(row=10, column=4), 5, bold=True, h='center', border=True)
    style(ws.cell(row=10, column=5), 5, bold=True, h='center', border=True)
    ws.merge_cells('F10:G10')
    style(ws.cell(row=10, column=6), 10, bold=True, h='center', border=True)
    style(ws.cell(row=10, column=7), '', border=True)
    style(ws.cell(row=10, column=8), 20, bold=True, h='center', border=True)
    style(ws.cell(row=10, column=9), 20, bold=True, h='center', border=True)
    style(ws.cell(row=10, column=10), 60, bold=True, h='center', border=True)
    style(ws.cell(row=10, column=11), '', border=True)
    style(ws.cell(row=10, column=12), '', border=True)

    # ROW 11: Column letters
    ws.row_dimensions[11].height = 18
    style(ws.cell(row=11, column=1), 'A', bold=True, h='center', border=True)
    ws.merge_cells('B11:C11')
    style(ws.cell(row=11, column=2), '', border=True)
    style(ws.cell(row=11, column=3), '', border=True)
    style(ws.cell(row=11, column=4), 'B', bold=True, h='center', border=True)
    style(ws.cell(row=11, column=5), 'C', bold=True, h='center', border=True)
    ws.merge_cells('F11:G11')
    style(ws.cell(row=11, column=6), 'D', bold=True, h='center', border=True)
    style(ws.cell(row=11, column=7), '', border=True)
    style(ws.cell(row=11, column=8), 'E', bold=True, h='center', border=True)
    style(ws.cell(row=11, column=9), 'F', bold=True, h='center', border=True)
    style(ws.cell(row=11, column=10), 'G', bold=True, h='center', border=True)
    style(ws.cell(row=11, column=11), 'H', bold=True, h='center', border=True)
    style(ws.cell(row=11, column=12), 'I', bold=True, h='center', border=True)

    # DATA ROWS
    marks_lookup = {m['traineeId']: m for m in subject_marks}
    current_row = 12
    for trainee in trainees:
        ws.row_dimensions[current_row].height = 18

        mark_entry = marks_lookup.get(trainee['id'], {})
        if subject_type == 'ES':
            target = mark_entry.get('totalESMarks', 15 if is_out_of_30 else 5)
        elif subject_type == 'WCS':
            target = mark_entry.get('totalWCSMarks', 5)
        else:
            target = mark_entry.get('totalEDMarks', 5)

        b, c, d, e, f, total_60 = distribute_marks(target, is_out_of_30)

        style(ws.cell(row=current_row, column=1),
              trainee.get('rollNumber') or trainee.get('enrollmentNumber', ''),
              h='left', border=True)

        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
        style(ws.cell(row=current_row, column=2), trainee.get('name', ''), h='left', border=True)
        style(ws.cell(row=current_row, column=3), '', border=True)

        style(ws.cell(row=current_row, column=4), b, h='center', border=True)
        style(ws.cell(row=current_row, column=5), c, h='center', border=True)

        ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=7)
        style(ws.cell(row=current_row, column=6), d, h='center', border=True)
        style(ws.cell(row=current_row, column=7), '', border=True)

        style(ws.cell(row=current_row, column=8), e, h='center', border=True)
        style(ws.cell(row=current_row, column=9), f, h='center', border=True)

        total_cell = ws.cell(row=current_row, column=10)
        total_cell.value = f"=SUM(D{current_row}:I{current_row})"
        total_cell.font = Font(name='Calibri', size=11)
        total_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_cell.border = thin_border()

        formula_suffix = '/2' if is_out_of_30 else '/6'
        conv_cell = ws.cell(row=current_row, column=11)
        conv_cell.value = f"=J{current_row}{formula_suffix}"
        conv_cell.font = Font(name='Calibri', size=11)
        conv_cell.alignment = Alignment(horizontal='center', vertical='center')
        conv_cell.border = thin_border()

        style(ws.cell(row=current_row, column=12), '', border=True)

        current_row += 1

    # Blank row before sign section
    current_row += 1

    # Sign of SI / Sign of FI line (merged B:K)
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=11)
    style(ws.cell(row=current_row, column=2),
          'Sign of SI :                                                            Sign of FI :')
    current_row += 1

    style(ws.cell(row=current_row, column=2), instructor.get('displayName', ''), bold=True)
    current_row += 1

    style(ws.cell(row=current_row, column=2), instructor.get('itiName', ''))
    style(ws.cell(row=current_row, column=8), instructor.get('itiName', ''))

    # Add thin outer border box around entire used range A1:L(last data area incl sign block)
    last_row = current_row
    thin = Side(style='thin')
    for r in range(1, last_row + 1):
        left_cell = ws.cell(row=r, column=1)
        right_cell = ws.cell(row=r, column=12)
        if r == 1:
            for cidx in range(1, 13):
                cc = ws.cell(row=r, column=cidx)
                cc.border = Border(top=thin, left=thin if cidx == 1 else cc.border.left,
                                    right=thin if cidx == 12 else cc.border.right,
                                    bottom=cc.border.bottom)
        if r == last_row:
            for cidx in range(1, 13):
                cc = ws.cell(row=r, column=cidx)
                cc.border = Border(bottom=thin, left=cc.border.left, right=cc.border.right, top=cc.border.top)
        lb = left_cell.border
        left_cell.border = Border(left=thin, top=lb.top, bottom=lb.bottom, right=lb.right)
        rb = right_cell.border
        right_cell.border = Border(right=thin, top=rb.top, bottom=rb.bottom, left=rb.left)

    wb.save(output_path)
    print(f'{subject_type} report saved: {output_path}')


if __name__ == '__main__':
    data_file = sys.argv[1]
    output_file = sys.argv[2]
    with open(data_file, 'r') as f:
        data = json.load(f)
    generate_subject_report(data, output_file)
