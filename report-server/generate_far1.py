import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def medium_border():
    m = Side(style='medium')
    return Border(left=m, right=m, top=m, bottom=m)

def thin_border():
    t = Side(style='thin')
    return Border(left=t, right=t, top=t, bottom=t)

# Criteria structure matching exact actual report
CRITERIA = [
    {'name': 'Safety consciousness', 'subs': [
        {'name': 'Dress code', 'max': 2},
        {'name': 'Use PPE', 'max': 5},
        {'name': 'Apply/practice safety', 'max': 8},
    ], 'total': 15},
    {'name': 'Workplace hygiene  & Economical use of materials', 'subs': [
        {'name': 'Maintain personal & \nworkplace cleanliness', 'max': 3},
        {'name': 'Dispose scrap as per \nstandard practice', 'max': 2},
        {'name': 'Select appropriate material \n&  minimize wastage', 'max': 5},
    ], 'total': 10},
    {'name': 'Attendance/ Punctuality', 'subs': [
        {'name': 'Initiative', 'max': 3},
        {'name': 'Accountability', 'max': 3},
        {'name': 'Participative in work', 'max': 4},
    ], 'total': 10},
    {'name': 'Ability to follow Manuals/ Written instructions', 'subs': [
        {'name': 'Select right manual', 'max': 1},
        {'name': 'Search for appropriate topic', 'max': 2},
        {'name': 'Read & interpret the manual', 'max': 2},
    ], 'total': 5},
    {'name': 'Application of Knowledge', 'subs': [
        {'name': 'Plan the work', 'max': 4},
        {'name': 'Select appropriate tools \n& equipment', 'max': 3},
        {'name': 'Review the work', 'max': 3},
    ], 'total': 10},
    {'name': 'Skills to handle tools & equipment', 'subs': [
        {'name': 'Handle & use tools & \nequipment', 'max': 4},
        {'name': 'Maintain safety in handling', 'max': 3},
        {'name': 'Care & maintain', 'max': 3},
    ], 'total': 10},
    {'name': 'Speed in doing work', 'subs': [
        {'name': 'Properly sequence the work', 'max': 3},
        {'name': 'Use appropriate technique', 'max': 5},
        {'name': 'Review the work during execution', 'max': 2},
    ], 'total': 10},
    {'name': 'Quality in workmanship', 'subs': [
        {'name': 'Achieve work with high accuracy', 'max': 7},
        {'name': 'Conform to requirement', 'max': 3},
        {'name': 'Satisfy the purpose', 'max': 5},
    ], 'total': 15},
    {'name': 'VIVA', 'subs': [
        {'name': 'Response with clarity', 'max': 7},
        {'name': 'Technical understanding', 'max': 5},
        {'name': 'Conscious towards job role', 'max': 3},
    ], 'total': 15},
]

# Exact column widths from actual report (col index -> width)
COL_WIDTHS = {
    1: 0.43, 2: 8.57, 3: 5.00,
    4: 3.57, 5: 2.86, 6: 3.14, 7: 3.43,
    8: 4.57, 9: 4.71, 10: 4.00, 11: 3.14,
    12: 3.43, 13: 2.86, 14: 3.43, 15: 3.57,
    16: 2.86, 17: 3.29, 18: 3.14, 19: 3.00,
    20: 3.14, 21: 4.29, 22: 3.29, 23: 3.57,
    24: 5.14, 25: 3.00, 26: 2.86, 27: 3.29,
    28: 8.43, 29: 4.86, 30: 4.57, 31: 3.86,
    32: 8.43, 33: 3.14, 34: 3.43, 35: 4.43,
    36: 3.00, 37: 8.43, 38: 3.14, 39: 8.43,
    40: 4.14, 41: 4.43, 42: 5.57, 43: 0.14,
}


def style_cell(cell, value='', bold=False, size=10, h='center', v='center',
               wrap=True, rotate=0, border=True, fill=None, font_color='FF000000'):
    cell.value = value
    cell.font = Font(name='Arial', size=size, bold=bold, color=font_color)
    cell.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap, text_rotation=rotate)
    if border:
        cell.border = medium_border()
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')


def generate_far1(data, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    trainees = data['trainees']
    distributed_marks = data['distributedMarks']
    instructor = data['instructorData']
    batch = data['batchData']
    half = data['half']
    assessment_date = data.get('assessmentDate', '')
    trade = data.get('tradeData', {}) or {}

    for trainee in trainees:
        trainee_id = trainee['id']
        trainee_marks = [m for m in distributed_marks
                          if m['traineeId'] == trainee_id and m['half'] == half]
        if not trainee_marks:
            continue

        lo_groups = {}
        for mark in trainee_marks:
            lo_id = mark['loId']
            if lo_id not in lo_groups:
                lo_groups[lo_id] = {
                    'loId': lo_id, 'loName': mark.get('loName', ''),
                    'loNumber': mark.get('loNumber', 0),
                    'loMark': mark.get('loMark', 0), 'practicals': []
                }
            lo_groups[lo_id]['practicals'].append(mark)
        sorted_los = sorted(lo_groups.values(), key=lambda x: x['loNumber'])

        sheet_name = (trainee.get('name') or f"T{trainee.get('enrollmentNumber','')}")[:28]
        for ch in ['/', '\\', '*', '?', '[', ']', ':']:
            sheet_name = sheet_name.replace(ch, '_')
        ws = wb.create_sheet(title=sheet_name or 'Sheet')

        ws.page_setup.orientation = 'landscape'
        ws.page_setup.paperSize = 9
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        for col_idx, w in COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w

        # Column A is a thin border spine (merged A1:A<lastrow>) - skip merging dynamically, just style
        for r in range(1, 9):
            style_cell(ws.cell(row=r, column=1), '', border=True)

        # ROW 1: Title — merged B1:AP1
        ws.row_dimensions[1].height = 15
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=42)
        style_cell(ws.cell(row=1, column=2), 'Internal Assessment', bold=True, size=11)
        for c in range(3, 43):
            style_cell(ws.cell(row=1, column=c), '')

        # ROW 2: Name of Trainee / Roll No / Year / Sem
        ws.row_dimensions[2].height = 15
        doa = trainee.get('dateOfAdmission', '')
        year_of_enrollment = batch.get('yearOfAssessment', '')
        if doa:
            if '/' in doa:
                parts = doa.split('/')
                if len(parts) == 3:
                    year_of_enrollment = parts[2]
            elif '-' in doa:
                year_of_enrollment = doa.split('-')[0]

        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=5)
        style_cell(ws.cell(row=2, column=2), 'Name of Trainee:', bold=False, h='left', border=False)
        ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=18)
        style_cell(ws.cell(row=2, column=6), trainee.get('name', ''), bold=True, h='left', border=False)
        style_cell(ws.cell(row=2, column=19), 'Roll NO:', bold=False, h='left', border=False)
        style_cell(ws.cell(row=2, column=22), trainee.get('rollNumber') or trainee.get('enrollmentNumber', ''), bold=True, border=False)
        ws.merge_cells(start_row=2, start_column=24, end_row=2, end_column=30)
        style_cell(ws.cell(row=2, column=24), 'Year of Enrollment:', bold=False, h='left', border=False)
        ws.merge_cells(start_row=2, start_column=31, end_row=2, end_column=34)
        style_cell(ws.cell(row=2, column=31), year_of_enrollment, bold=True, border=False)
        style_cell(ws.cell(row=2, column=35), 'Sem:', bold=False, h='left', border=False)
        style_cell(ws.cell(row=2, column=39), half, bold=True, border=False)

        # ROW 3: Name of ITI / Date of Assessment / Batch
        ws.row_dimensions[3].height = 15
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=5)
        style_cell(ws.cell(row=3, column=2), 'Name of ITI:', bold=False, h='left', border=False)
        ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=18)
        style_cell(ws.cell(row=3, column=6), instructor.get('itiName', ''), bold=True, h='left', border=False)
        ws.merge_cells(start_row=3, start_column=24, end_row=3, end_column=30)
        style_cell(ws.cell(row=3, column=24), 'Date of Assessment:', bold=False, h='left', border=False)
        ws.merge_cells(start_row=3, start_column=31, end_row=3, end_column=34)
        style_cell(ws.cell(row=3, column=31), assessment_date, bold=True, border=False)
        style_cell(ws.cell(row=3, column=35), 'Batch:', bold=False, h='left', border=False)
        style_cell(ws.cell(row=3, column=39), batch.get('batchNumber', ''), bold=True, border=False)

        # ROW 4: Name of Industry / Assessment Location
        ws.row_dimensions[4].height = 15
        ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)
        style_cell(ws.cell(row=4, column=2), 'Name of the Industry:', bold=False, h='left', border=False)
        ws.merge_cells(start_row=4, start_column=6, end_row=4, end_column=23)
        style_cell(ws.cell(row=4, column=6), trade.get('name', ''), bold=True, h='left', border=False)
        ws.merge_cells(start_row=4, start_column=24, end_row=4, end_column=30)
        style_cell(ws.cell(row=4, column=24), 'Assessment Location:', bold=False, h='left', border=False)
        ws.merge_cells(start_row=4, start_column=31, end_row=4, end_column=42)
        style_cell(ws.cell(row=4, column=31), instructor.get('address', ''), bold=True, border=False)

        # ROW 5: Trade Name / Duration / S.I.Name
        ws.row_dimensions[5].height = 15
        duration = trade.get('duration', 1)
        ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=5)
        style_cell(ws.cell(row=5, column=2), 'Trade Name:', bold=False, h='left', border=False)
        ws.merge_cells(start_row=5, start_column=6, end_row=5, end_column=23)
        style_cell(ws.cell(row=5, column=6), trade.get('name', ''), bold=True, h='left', border=False)
        ws.merge_cells(start_row=5, start_column=24, end_row=5, end_column=30)
        style_cell(ws.cell(row=5, column=24), 'Duration of the Trade:', bold=False, h='left', border=False)
        ws.merge_cells(start_row=5, start_column=31, end_row=5, end_column=34)
        style_cell(ws.cell(row=5, column=31), f"{duration} Year", bold=True, border=False)
        style_cell(ws.cell(row=5, column=35), 'S.I.Name:', bold=False, h='left', border=False)
        ws.merge_cells(start_row=5, start_column=38, end_row=5, end_column=42)
        style_cell(ws.cell(row=5, column=38), instructor.get('displayName', ''), bold=True, border=False)

        # ROW 6: Criteria group headers (merged across sub+total cols)
        ws.row_dimensions[6].height = 42.75
        style_cell(ws.cell(row=6, column=2), '')
        style_cell(ws.cell(row=6, column=3), '')
        col = 4
        for criteria in CRITERIA:
            start_col = col
            end_col = col + len(criteria['subs'])  # includes total col
            ws.merge_cells(start_row=6, start_column=start_col, end_row=6, end_column=end_col)
            style_cell(ws.cell(row=6, column=start_col), criteria['name'], bold=True, size=10, wrap=True)
            for cc in range(start_col + 1, end_col + 1):
                style_cell(ws.cell(row=6, column=cc), '')
            col = end_col + 1
        style_cell(ws.cell(row=6, column=40), '')
        style_cell(ws.cell(row=6, column=41), '')
        style_cell(ws.cell(row=6, column=42), '')

        # ROW 7: Sub-criteria headers (rotated 90)
        ws.row_dimensions[7].height = 126.0
        style_cell(ws.cell(row=7, column=2), 'Learning Outcome Number', bold=True, size=8, rotate=90)
        style_cell(ws.cell(row=7, column=3), 'Practical / \nProfessional Skill Number', bold=True, size=8, rotate=90)
        col = 4
        for criteria in CRITERIA:
            for sub in criteria['subs']:
                style_cell(ws.cell(row=7, column=col), sub['name'], bold=True, size=8, rotate=90)
                col += 1
            style_cell(ws.cell(row=7, column=col), 'Total', bold=True, size=8, rotate=90)
            col += 1
        style_cell(ws.cell(row=7, column=40), 'Grand Total', bold=True, size=8, rotate=90)
        style_cell(ws.cell(row=7, column=41), 'Signature of Trainee', bold=True, size=8, rotate=90)
        style_cell(ws.cell(row=7, column=42), 'Signature of SI', bold=True, size=8, rotate=90)

        # ROW 8: Max marks
        ws.row_dimensions[8].height = 15
        style_cell(ws.cell(row=8, column=2), '')
        style_cell(ws.cell(row=8, column=3), '')
        col = 4
        for criteria in CRITERIA:
            for sub in criteria['subs']:
                style_cell(ws.cell(row=8, column=col), str(sub['max']), bold=True)
                col += 1
            style_cell(ws.cell(row=8, column=col), str(criteria['total']), bold=True)
            col += 1
        style_cell(ws.cell(row=8, column=40), '100', bold=True)
        style_cell(ws.cell(row=8, column=41), '')
        style_cell(ws.cell(row=8, column=42), '')

        # DATA ROWS
        current_row = 9
        for lo in sorted_los:
            sorted_practicals = sorted(lo['practicals'], key=lambda x: x.get('practicalNumber', 0))

            for practical in sorted_practicals:
                ws.row_dimensions[current_row].height = 15
                style_cell(ws.cell(row=current_row, column=1), '')
                style_cell(ws.cell(row=current_row, column=2), f"LO - {lo['loNumber']}")
                style_cell(ws.cell(row=current_row, column=3), practical.get('practicalNumber', ''))

                data_col = 4
                grand_total = 0
                for criteria in practical.get('criteriaMarks', []):
                    for sub_mark in criteria.get('subCriteriaMarks', []):
                        val = sub_mark.get('allocatedMark', 0)
                        style_cell(ws.cell(row=current_row, column=data_col), val)
                        data_col += 1
                    c_total = criteria.get('allocatedMark', 0)
                    style_cell(ws.cell(row=current_row, column=data_col), c_total)
                    grand_total += c_total
                    data_col += 1

                style_cell(ws.cell(row=current_row, column=40), grand_total)
                style_cell(ws.cell(row=current_row, column=41), '')
                style_cell(ws.cell(row=current_row, column=42), '')
                current_row += 1

            # LO average row — light cyan fill, label B:AF, "Average of LOx" AG:AM, value AN
            ws.row_dimensions[current_row].height = 15
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=32)
            style_cell(ws.cell(row=current_row, column=2),
                       lo.get('loName', f"LO {lo['loNumber']}"),
                       bold=True, h='left', fill='FFAFEEEE')
            for c in range(3, 33):
                style_cell(ws.cell(row=current_row, column=c), '', fill='FFAFEEEE')

            ws.merge_cells(start_row=current_row, start_column=33, end_row=current_row, end_column=39)
            style_cell(ws.cell(row=current_row, column=33),
                       f'Average of LO{lo["loNumber"]}', bold=True, h='right', fill='FFAFEEEE')
            for c in range(34, 40):
                style_cell(ws.cell(row=current_row, column=c), '', fill='FFAFEEEE')

            style_cell(ws.cell(row=current_row, column=40), lo.get('loMark', 0),
                       bold=True, fill='FFAFEEEE')
            style_cell(ws.cell(row=current_row, column=41), '', fill='FFAFEEEE')
            style_cell(ws.cell(row=current_row, column=42), '', fill='FFAFEEEE')

            current_row += 1

        # Overall average row — gray fill, B:AM merged, value AN
        overall_avg = round(sum(lo.get('loMark', 0) for lo in sorted_los) / len(sorted_los)) if sorted_los else 0
        ws.row_dimensions[current_row].height = 15
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=39)
        style_cell(ws.cell(row=current_row, column=2), 'Average of All LO',
                   bold=True, h='left', fill='FFD3D3D3')
        for c in range(3, 40):
            style_cell(ws.cell(row=current_row, column=c), '', fill='FFD3D3D3')
        style_cell(ws.cell(row=current_row, column=40), overall_avg, bold=True, fill='FFD3D3D3')
        style_cell(ws.cell(row=current_row, column=41), '', fill='FFD3D3D3')
        style_cell(ws.cell(row=current_row, column=42), '', fill='FFD3D3D3')

        # Column A thin spine border for full sheet height
        ws.merge_cells(start_row=1, start_column=1, end_row=current_row, end_column=1)

    wb.save(output_path)
    print(f"FAR-1 saved: {output_path}")


if __name__ == '__main__':
    data_file = sys.argv[1]
    output_file = sys.argv[2]
    with open(data_file, 'r') as f:
        data = json.load(f)
    generate_far1(data, output_file)